from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

#create a workbook obj:
wb=Workbook()

#create an active worksheet:
ws=wb.active

#loadexisting spreadsheet

wb=load_workbook('Book1.xlsx')
