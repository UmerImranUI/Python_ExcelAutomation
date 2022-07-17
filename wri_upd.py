from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

# #create a workbook obj:
# wb=Workbook()
wb=load_workbook('Book1.xlsx')

ws=wb['Sheet1']


#update
# ws['B5'].value='cute'
# ws['A9'].value='kayan'
# ws['B9'].value='magenda'
# ws.cell(row=5, column=2).value=5000
# print(ws.cell(row=5, column=2).value)
# wb.save('Book1.xlsx')
ws['D1']='Double Degrees'
for i in range(2,10):
    c_col=ws.cell(row=i, column=3).value
    d_value=c_col*2
    ws.cell(row=i, column=4).value=d_value

wb.save('Book1.xlsx')
