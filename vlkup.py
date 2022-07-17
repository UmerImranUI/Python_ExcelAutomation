from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
  

# #create a workbook obj:
# wb=Workbook()
master_data=load_workbook('MOCK_DATA.xlsx')
daily_data=load_workbook('MOCK_DATA2.xlsx')

master_sheet=master_data['data']
daily_sheet=daily_data['data']

#logic to print the sheet's data
val=[]
for i in range(2,5):
    val.append(master_sheet.cell(row=1, column=i).value)
i=0
for j in range(4,7):
    
    daily_sheet.cell(row=1, column=j).value=val[i]
    i=i+1
    



for i in daily_sheet.iter_rows():
    id=i[0].value       #prints row's first value of id
    row_number=i[0].row
    # print(id)
    # print(row_number)  #prints row num
    for j in master_sheet.iter_rows():
        if j[0].value==id:
            
            daily_sheet.cell(row=row_number, column=4).value=j[1].value
            daily_sheet.cell(row=row_number, column=5).value=j[2].value
            daily_sheet.cell(row=row_number, column=6).value=j[3].value
            
daily_data.save('updated_sheet.xlsx')
