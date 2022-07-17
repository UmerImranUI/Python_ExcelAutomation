import openpyxl
from openpyxl.styles import numbers

wb=openpyxl.load_workbook('Book1.xlsx')
ws=wb.active

ws['C4']='11/11/20'
ws['C4'].number_format = numbers.FORMAT_DATE_DATETIME

ws['D4']=20

ws['E4']='Beginner'
ws['E4'].number_format = numbers.FORMAT_TEXT

wb.save('Book1.xlsx')