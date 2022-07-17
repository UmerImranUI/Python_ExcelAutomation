from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
  

# #create a workbook obj:
# wb=Workbook()
data=load_workbook('sprecial char.xlsx')
names=data['Sheet1']

is_data=True
row_count=1

while is_data:
    row_count+=1
    first_name = names.cell(row=row_count, column=1).value

    if first_name:
        names.cell(row=row_count, column=1).value=first_name.strip('"')
    else:
        is_data=False

data.save('outputchar.xlsx')
