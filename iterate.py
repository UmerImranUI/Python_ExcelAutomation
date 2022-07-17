from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

# #create a workbook obj:
# wb=Workbook()
wb=load_workbook('Book1.xlsx')

#create an active worksheet:
ws=wb.active

rows=ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2)
print(rows)

# for a, b in rows:
#     print(a.value, b.value)

# names=[]
# colors=[]
# for a, b in rows:
#     names.append(a.value)
#     colors.append(b.value)

# print(names)
# print(colors)

# columns=ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=2)
columns=ws.iter_cols(max_row=5, max_col=2)
#these functions give all the values if we dont specify min col, rows
for col in columns:
    print(col)


cols=list(ws.columns)
print(cols)