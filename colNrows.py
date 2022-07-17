from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

# #create a workbook obj:
# wb=Workbook()
wb = load_workbook('Book1.xlsx')

#create an active worksheet:
ws = wb.active

# column_a = ws['A']
column_a = ws['1'] #for 1st row
# column_b = ws['B']
print(ws.cell(row=6, column=2).value)


for cell in column_a:
    print(cell.value)
# for cell in column_b:
#     print(cell.value)