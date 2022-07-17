from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

# #create a workbook obj:
# wb=Workbook()
wb = load_workbook('Book1.xlsx')

#create an active worksheet:
ws = wb.active

fill_pattern=PatternFill(patternType='solid', fgColor='C64949')
ws['B4'].fill=fill_pattern
wb.save('Book1.xlsx')