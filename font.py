from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color

# #create a workbook obj:
# wb=Workbook()
wb = load_workbook('Book1.xlsx')

#create an active worksheet:
ws = wb.active
# font_style=Font(name='Chalkboard', size='14', color='1A4FDF', italic=True, bold=True)
# #font styles changing
# a4=ws['A4']
# a4.font=font_style
# wb.save('Book1.xlsx')

col_style=font_style=Font(name='Chalkboard', size='14', color='1A4FDF', italic=True, bold=True, underline='single')

for i in range(2, 10):
    ws.cell(row=i, column=3).font=col_style


wb.save('Book1.xlsx')



