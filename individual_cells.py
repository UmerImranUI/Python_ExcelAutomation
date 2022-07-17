from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

# #create a workbook obj:
# wb=Workbook()
wb=load_workbook('Book1.xlsx')

#create an active worksheet:
ws=wb.active

#print cell from the spreadsheet
# print(ws['A2'].value)

# print(f'{ws["A2"].value}: {ws["B2"].value}')

#by var
name=ws["A2"].value
color=ws["B2"].value

print(f'{name}: {color}')

