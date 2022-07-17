from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl  

# #create a workbook obj:
# wb=Workbook()
master_data=load_workbook('MOCK_DATA.xlsx')
daily_data=load_workbook('MOCK_DATA2.xlsx')

master_sheet=master_data['data']
daily_sheet=daily_data['data']


#rows count of daily sheet
is_data=True
daily_row_count=1

while is_data:
    daily_row_count += 1
    data=daily_sheet.cell(row=daily_row_count, column=1).value
    if data == None:
        is_data=False
# print(daily_row_count)



is_data=True
master_row_count=1

while is_data:
    master_row_count += 1
    data=master_sheet.cell(row=master_row_count, column=1).value
    if data == None:
        is_data=False
# print(master_row_count)

#get data from daily_sheet
#extract data --> store into list of dict

todays_data=[]
for i in range(1, daily_row_count):
    row_data={}
    row_data['id']=daily_sheet.cell(row=i, column=1).value
    row_data['todays_purchase']=daily_sheet.cell(row=i, column=2).value
    row_data['todays_rewards']=daily_sheet.cell(row=i, column=3).value
    todays_data.append(row_data)

# print(todays_data)

#{'id': 'id', 'total purchase': 'Total purchase', 'Total Reward': 'Total Reward'}

# write daily sheet data into master excel sheet
# Find row using the ID 
# Go to total purchase cell+todays purchase
# Go to total reward balance + todays reward 

for i in range(2, master_row_count):
    id = master_sheet.cell(row=i, column=1).value
    for row in todays_data:
        if row['id'] == id:
            todays_purchase=row['todays_purchase']
            todays_reward=row['todays_rewards']

            #get data from master sheet
            total_purchase=master_sheet.cell(row=i, column=6).value
            total_reward=master_sheet.cell(row=i, column=7).value

            #Add values of todays data into total data
            new_total_purchase=total_purchase+todays_purchase
            new_total_reward=total_reward+todays_reward

            master_sheet.cell(row=i, column=6).value = new_total_purchase
            master_sheet.cell(row=i, column=7).value = new_total_reward

# master_data.save('MOCK_DATA_UPDATED.xlsx')


daily_report=openpyxl.Workbook()
ws=daily_report.active

#get headers
is_data=True
column_count=1
header_values=[]

while is_data:
    column_count+=1
    data = master_sheet.cell(row=1, column=column_count).value
    if data != None:
        header_values.append(data)
    else:
        is_data=False
header_style=Font(name="Times New Roman", size=12, bold=True)
# print(header_values)

for i, col_name in enumerate(header_values):
    col_index = i+1
    ws.cell(row=1, column=col_index).value=col_name
    ws.cell(row=1, column=col_index).font=header_style
    
IDs=[]
for data in todays_data:
    IDs.append(data['id'])
IDs.pop(0)
# print(IDs)


final_data=[]
for i in range(2, master_row_count):
    id=master_sheet.cell(row=i, column=1).value
    if id in IDs:
        lst=[]
        for j in range(2,8):
            lst.append(master_sheet.cell(row=i, column=j).value)
        final_data.append(lst)
        

for data in final_data:
    ws.append(data)

daily_report.save("daily_report_send.xlsx")