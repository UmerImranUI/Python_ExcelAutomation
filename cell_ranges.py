from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook

# #create a workbook obj:
# wb=Workbook()
wb = load_workbook('Book1.xlsx')

#create an active worksheet:
ws = wb.active

#grab a range
range= ws['A2':'A10']

# print(range)

for cell in range:  #accessing the range
    for cells in cell:
        print(cells.value)