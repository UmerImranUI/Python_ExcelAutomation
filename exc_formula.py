import openpyxl
from openpyxl.styles import Font

wb=openpyxl.load_workbook('Book1.xlsx')
ws=wb.active

ws['B11']="=SUM(B2:B9)"
ws['B12']="=AVERAGE(B2:B9)"

#(balance*Interest rate)+balance

ws['D1']='Balance after a year'
ws['D1'].font=Font(bold=True, name='Arial')

for i in range(2,10):
    balance=ws.cell(row=i, column=2).value  
    interest=ws.cell(row=i, column=3).value  
    finaL_balance=(balance*interest)+balance
    ws.cell(row=i, column=4).value=finaL_balance

wb.save("Book1.xlsx")