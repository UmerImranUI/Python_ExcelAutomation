from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

# #create a workbook obj:
# wb=Workbook()
wb = load_workbook('Book1.xlsx')

#create an active worksheet:
ws = wb.active

top=Side(border_style='dashed', color='FF0707')
bottom=Side(border_style='double', color='FF0707')
border=Border(top=top, bottom=bottom)

ws['B6'].border=border

wb.save('Book1.xlsx')